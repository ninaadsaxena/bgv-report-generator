#!/usr/bin/env python3
"""
BGV Report Generator — Flask API Backend  (v2.2)
=================================================
Changes v2.2:
  - Uploads stored in-memory as bytes — nothing written to disk on upload
  - Generate job writes a short-lived temp file only during Word COM conversion,
    then deletes it immediately — zero disk footprint at rest
  - PDFs stored in-memory + persisted as base64 in history.json
  - No UPLOADS_DIR or OUTPUT_DIR on disk
"""

import os, uuid, json, io, base64, shutil, threading, tempfile, zipfile
from pathlib import Path
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, send_file, send_from_directory, Response
from flask_cors import CORS

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BASE_DIR      = Path(__file__).parent.resolve()
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR    = BASE_DIR / "static"
HISTORY_FILE  = BASE_DIR / "history.json"

for d in (TEMPLATES_DIR, STATIC_DIR):
    d.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
CORS(app)

# ---------------------------------------------------------------------------
# In-memory stores
# ---------------------------------------------------------------------------

jobs: dict[str, dict] = {}
jobs_lock = threading.Lock()

# uploads[upload_id] = {original_filename, data_b64, sheets, data, uploaded_at}
uploads: dict[str, dict] = {}
uploads_lock = threading.Lock()


# ---------------------------------------------------------------------------
# History helpers  (stores PDF bytes as base64 for cross-restart access)
# ---------------------------------------------------------------------------

def load_history() -> list:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def save_history(history: list):
    HISTORY_FILE.write_text(json.dumps(history, indent=2, default=str), encoding="utf-8")


def append_history(entry: dict):
    history = load_history()
    history.insert(0, entry)
    history = history[:200]
    save_history(history)


# ---------------------------------------------------------------------------
# Cleanup: drop in-memory uploads older than 2h
# ---------------------------------------------------------------------------

def cleanup_old_uploads():
    cutoff = datetime.now() - timedelta(hours=2)
    with uploads_lock:
        expired = [uid for uid, u in uploads.items()
                   if datetime.fromisoformat(u["uploaded_at"]) < cutoff]
        for uid in expired:
            del uploads[uid]


# ---------------------------------------------------------------------------
# Routes — frontend
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


# ---------------------------------------------------------------------------
# Routes — Templates
# ---------------------------------------------------------------------------

@app.route("/api/templates", methods=["GET"])
def list_templates():
    templates = []
    for f in TEMPLATES_DIR.iterdir():
        if f.suffix.lower() == ".docx" and f.is_file():
            templates.append({
                "name":     f.name,
                "size":     f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            })
    return jsonify({"templates": sorted(templates, key=lambda x: x["name"])})


@app.route("/api/templates/upload", methods=["POST"])
def upload_template():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".docx"):
        return jsonify({"error": "Only .docx files are accepted"}), 400
    save_path = TEMPLATES_DIR / f.filename
    f.save(str(save_path))
    return jsonify({"message": f"Template '{f.filename}' uploaded", "name": f.filename})


@app.route("/api/templates/<name>", methods=["DELETE"])
def delete_template(name):
    path = TEMPLATES_DIR / name
    if not path.exists():
        return jsonify({"error": "Template not found"}), 404
    path.unlink()
    return jsonify({"message": f"Template '{name}' deleted"})


# ---------------------------------------------------------------------------
# Routes — Excel upload & preview
# ---------------------------------------------------------------------------

@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted"}), 400

    cleanup_old_uploads()   # evict old in-memory uploads

    file_bytes = f.read()   # read entirely into memory — nothing touches disk
    upload_id  = str(uuid.uuid4())

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets = wb.sheetnames
        result_sheets = {}

        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result_sheets[sheet_name] = {"headers": [], "rows": [], "row_count": 0}
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            data_rows = []
            for raw in rows[1:]:
                if raw is None or all(v is None for v in raw):
                    continue
                serialized = {}
                for h, v in zip(headers, raw):
                    serialized[h] = v.isoformat() if hasattr(v, "isoformat") else ("" if v is None else str(v))
                data_rows.append(serialized)
            result_sheets[sheet_name] = {
                "headers":   headers,
                "rows":      data_rows[:50],
                "row_count": len(data_rows),
            }

        # Store in-memory — base64 so JSON-serialisable if ever needed
        with uploads_lock:
            uploads[upload_id] = {
                "original_filename": f.filename,
                "data_b64":          base64.b64encode(file_bytes).decode("ascii"),
                "sheets":            sheets,
                "data":              result_sheets,
                "uploaded_at":       datetime.now().isoformat(),
            }

        return jsonify({
            "upload_id":         upload_id,
            "original_filename": f.filename,
            "sheets":            sheets,
            "data":              result_sheets,
        })
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel file: {e}"}), 400


# ---------------------------------------------------------------------------
# Routes — Generate
# ---------------------------------------------------------------------------

@app.route("/api/generate", methods=["POST"])
def generate():
    data = request.get_json(force=True)
    upload_id     = data.get("upload_id")
    sheet_name    = data.get("sheet_name") or None
    template_name = data.get("template_name") or None

    with uploads_lock:
        upload_entry = uploads.get(upload_id)
    if not upload_entry:
        return jsonify({"error": "Upload not found — please re-upload your file"}), 404

    original_filename = upload_entry["original_filename"]
    excel_bytes = base64.b64decode(upload_entry["data_b64"])

    template_override = None
    if template_name:
        tp = TEMPLATES_DIR / template_name
        if not tp.exists():
            return jsonify({"error": f"Template '{template_name}' not found"}), 404
        template_override = str(tp)

    job_id = str(uuid.uuid4())
    with jobs_lock:
        jobs[job_id] = {
            "status":            "running",
            "progress":          0,
            "total":             0,
            "messages":          [],
            "pdfs":              [],   # [{name, data_b64}]
            "report_meta":       None,
            "started_at":        datetime.now().isoformat(),
            "original_filename": original_filename,
        }

    def run_job():
        from generate_reports import generate_all

        def on_progress(done, total, msg):
            with jobs_lock:
                jobs[job_id]["progress"] = done
                jobs[job_id]["total"]    = max(total, 1)
                jobs[job_id]["messages"].append(msg)

        try:
            # Write the Excel bytes to a temp file for the duration of this job only
            ext = Path(original_filename).suffix or ".xlsx"
            tmp_xlsx = Path(tempfile.gettempdir()) / f"bgv_tmp_{job_id}{ext}"
            try:
                tmp_xlsx.write_bytes(excel_bytes)

                report = generate_all(
                    str(tmp_xlsx),
                    sheet_name=sheet_name,
                    template_override=template_override,
                    progress_callback=on_progress,
                )
            finally:
                tmp_xlsx.unlink(missing_ok=True)   # delete immediately after use

            # Convert PDF bytes to base64 for storage
            pdfs_meta = []
            for g in report["generated"]:
                pdfs_meta.append({
                    "filename":   g["filename"],
                    "applicant":  g["applicant"],
                    "check_name": g["check_name"],
                    "data_b64":   base64.b64encode(g["pdf_bytes"]).decode("ascii"),
                    "warnings":   g["warnings"],
                    "row":        g["row"],
                })

            report_meta = {
                "generated_count": len(report["generated"]),
                "error_count":     len(report["errors"]),
                "errors":          report["errors"],
            }

            with jobs_lock:
                jobs[job_id]["status"]      = "done"
                jobs[job_id]["pdfs"]        = pdfs_meta
                jobs[job_id]["report_meta"] = report_meta
                jobs[job_id]["progress"]    = jobs[job_id]["total"]

            # Persist to history (include PDF data so downloads survive restarts)
            append_history({
                "job_id":            job_id,
                "original_filename": original_filename,
                "started_at":        jobs[job_id]["started_at"],
                "completed_at":      datetime.now().isoformat(),
                "pdfs":              pdfs_meta,        # includes data_b64
                "error_count":       len(report["errors"]),
                "errors":            report["errors"],
            })

        except Exception as e:
            with jobs_lock:
                jobs[job_id]["status"]      = "error"
                jobs[job_id]["messages"].append(f"Fatal: {e}")
                jobs[job_id]["report_meta"] = {"generated_count": 0, "error_count": 1, "errors": [str(e)]}


    threading.Thread(target=run_job, daemon=True).start()
    return jsonify({"job_id": job_id})


# ---------------------------------------------------------------------------
# Routes — Status
# ---------------------------------------------------------------------------

@app.route("/api/status/<job_id>", methods=["GET"])
def job_status(job_id):
    with jobs_lock:
        job = jobs.get(job_id)

    if job is None:
        # Attempt to reconstruct from history
        history = load_history()
        entry = next((h for h in history if h["job_id"] == job_id), None)
        if entry:
            return jsonify({
                "status":      "done",
                "progress":    len(entry.get("pdfs", [])),
                "total":       len(entry.get("pdfs", [])),
                "messages":    [],
                "report_meta": {
                    "generated_count": len(entry.get("pdfs", [])),
                    "error_count":     entry.get("error_count", 0),
                    "errors":          entry.get("errors", []),
                },
                "pdfs": [{"filename": p["filename"], "applicant": p["applicant"],
                           "check_name": p["check_name"], "warnings": p.get("warnings", [])}
                          for p in entry.get("pdfs", [])],
            })
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status":      job["status"],
        "progress":    job["progress"],
        "total":       job["total"],
        "messages":    job["messages"][-20:],
        "report_meta": job.get("report_meta"),
        "pdfs": [{"filename": p["filename"], "applicant": p["applicant"],
                   "check_name": p["check_name"], "warnings": p.get("warnings", [])}
                  for p in job.get("pdfs", [])],
    })


# ---------------------------------------------------------------------------
# Routes — Download (FIXED: filename embedded in URL path)
# ---------------------------------------------------------------------------

def _get_pdf_bytes(job_id, filename):
    """Look up PDF bytes from in-memory jobs or from history."""
    with jobs_lock:
        job = jobs.get(job_id)
    if job:
        for p in job.get("pdfs", []):
            if p["filename"] == filename:
                return base64.b64decode(p["data_b64"])

    # Fallback: history file
    history = load_history()
    entry = next((h for h in history if h["job_id"] == job_id), None)
    if entry:
        for p in entry.get("pdfs", []):
            if p["filename"] == filename:
                return base64.b64decode(p["data_b64"])
    return None


@app.route("/api/download/<job_id>/<path:filename>", methods=["GET"])
def download_pdf(job_id, filename):
    """Download a single PDF — filename is part of URL so browser uses it natively."""
    pdf_bytes = _get_pdf_bytes(job_id, filename)
    if pdf_bytes is None:
        return jsonify({"error": "File not found or job expired"}), 404

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        }
    )


@app.route("/api/download/<job_id>", methods=["GET"])
def download_all_pdfs(job_id):
    """
    Download all PDFs from a job.
    Single PDF → serves directly as PDF.
    Multiple PDFs → serves as ZIP.
    """
    with jobs_lock:
        job = jobs.get(job_id)

    pdfs = []
    if job:
        pdfs = job.get("pdfs", [])
    else:
        history = load_history()
        entry = next((h for h in history if h["job_id"] == job_id), None)
        if entry:
            pdfs = entry.get("pdfs", [])

    if not pdfs:
        return jsonify({"error": "No PDFs found for this job"}), 404

    if len(pdfs) == 1:
        # Single — redirect to the clean named URL so browser uses the proper filename
        from flask import redirect
        return redirect(f"/api/download/{job_id}/{pdfs[0]['filename']}")

    # Multiple — bundle as ZIP
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in pdfs:
            zf.writestr(p["filename"], base64.b64decode(p["data_b64"]))
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"BGV_Reports_{job_id[:8]}.zip",
        mimetype="application/zip",
    )


# ---------------------------------------------------------------------------
# Routes — History
# ---------------------------------------------------------------------------

@app.route("/api/history", methods=["GET"])
def get_history():
    history = load_history()
    # Return metadata only (strip data_b64 from pdfs for speed)
    slim = []
    for entry in history:
        slim.append({
            "job_id":            entry["job_id"],
            "original_filename": entry.get("original_filename", ""),
            "started_at":        entry.get("started_at", ""),
            "completed_at":      entry.get("completed_at", ""),
            "error_count":       entry.get("error_count", 0),
            "pdfs": [
                {"filename": p["filename"], "applicant": p["applicant"],
                 "check_name": p["check_name"]}
                for p in entry.get("pdfs", [])
            ],
        })
    return jsonify({"history": slim})


@app.route("/api/history/<job_id>", methods=["DELETE"])
def delete_history_entry(job_id):
    history = load_history()
    if not any(h["job_id"] == job_id for h in history):
        return jsonify({"error": "Entry not found"}), 404
    history = [h for h in history if h["job_id"] != job_id]
    save_history(history)
    with jobs_lock:
        jobs.pop(job_id, None)
    return jsonify({"message": "Deleted"})


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 60)
    print("  BGV Report Generator  v2.1")
    print("  Open http://localhost:5000 in your browser")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
